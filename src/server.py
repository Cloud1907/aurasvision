"""VideoAI web sunucusu — FastAPI.

Basit, kullanıcı-dostu yönetim arayüzü için backend:
  GET  /                     → web UI (web/index.html)
  GET  /api/cameras          → config'teki kameralar
  GET  /api/snapshot?camera= → kameradan anlık kare (JPEG; diske YAZILMAZ — KVKK)
  GET  /api/zones?camera=    → bölge/çizgi tanımları
  POST /api/zones            → bölge/çizgi kaydet (canvas editöründen)
  GET  /api/events?limit=    → birleşik olay akışı (count/plate/face)
  GET  /api/lists?kind=      → izleme listesi (plate | face)
  POST /api/lists            → listeye ekle
  DELETE /api/lists/{kind}/{id} → listeden sil

Çalıştırma: python -m src.server   (veya uvicorn src.server:app)
"""
from __future__ import annotations

import json
import logging
import os
import secrets
import threading
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path

import cv2
from fastapi import (FastAPI, HTTPException, Query, Request, WebSocket,
                     WebSocketDisconnect)
from fastapi.responses import FileResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from . import akis, kimlik
from .config import apply_cv2_http_headers, http_options, load_config
from .store import DEFAULT_TASKS, merged_cameras, open_store

cfg = load_config()
apply_cv2_http_headers(cfg)   # cv2 tabanlı analiz de CDN başlıklarını göndersin
ROOT = Path(__file__).resolve().parent.parent
WEB_DIR = ROOT / "web"

app = FastAPI(title="AurasVision")

# Substream go2rtc'de ayrı bir akış olarak yayınlanır: "<kamera-id><SUB_SUFFIX>"
SUB_SUFFIX = "-sub"

# Erişim modeli (docs/rbac-tasarim.md):
#   1) AURAS_TOKEN — makine anahtarı (script, exporter, mobil QR). Geriye uyum:
#      her zaman geçerli, yonetici yetkisinde.
#   2) Kullanıcı/rol — kullanicilar tablosu doluysa parola girişi + imzalı çerez;
#      yazma uçları rol matrisinden (kimlik.yetkili) geçer.
#   3) İkisi de yoksa auth kapalı (yerel demo) — eski davranış.
API_TOKEN = os.environ.get("AURAS_TOKEN", "")

# Kullanıcılar her istekte DB'ye sorulmaz — 10 sn önbellek; kullanıcı
# ekleme/silme ucu önbelleği anında düşürür (_kullanici_cache_sifirla)
_KUL_CACHE: dict = {"roller": None, "t": 0.0}


def _kullanici_rolleri() -> dict:
    """ad → rol sözlüğü (10 sn önbellek).

    Çerez yalnız KİMLİĞİ kanıtlar; ROL buradan okunur. Aksi hâlde silinen
    kullanıcının çerezi 12 saat geçerli kalır, rol düşürme de oturum bitene
    dek işlemezdi — denetim açısından kabul edilemez.
    """
    if _KUL_CACHE["roller"] is None or time.time() - _KUL_CACHE["t"] > 10:
        s = _store()
        try:
            _KUL_CACHE["roller"] = {u["ad"]: u["rol"] for u in s.kullanici_listele()}
        except Exception:
            _KUL_CACHE["roller"] = {}
        finally:
            s.close()
        _KUL_CACHE["t"] = time.time()
    return _KUL_CACHE["roller"]


def _kullanici_var() -> bool:
    return bool(_kullanici_rolleri())


def _kullanici_cache_sifirla() -> None:
    _KUL_CACHE["roller"] = None


@app.middleware("http")
async def _auth(request: Request, call_next):
    p = request.url.path
    request.state.kullanici = None
    if not (p.startswith("/api") or p.startswith("/media")):
        return await call_next(request)
    if p in ("/api/giris", "/api/health"):
        return await call_next(request)   # giriş kapısı ve canlılık her zaman açık
    tok = request.headers.get("authorization", "")
    tok = tok[7:] if tok.lower().startswith("bearer ") else request.query_params.get("token", "")
    if API_TOKEN and tok and secrets.compare_digest(tok, API_TOKEN):
        request.state.kullanici = {"ad": "sistem", "rol": "yonetici"}
        return await call_next(request)
    oturum = kimlik.coz(request.cookies.get(kimlik.OTURUM_CEREZ, ""))
    if oturum:
        rol = _kullanici_rolleri().get(oturum["ad"])
        if rol is None:
            oturum = None   # kullanıcı silinmiş → çerez artık kimlik değil
        else:
            oturum["rol"] = rol   # rol değişikliği yeni oturum beklemeden işler
            request.state.kullanici = oturum
            if not kimlik.yetkili(rol, request.method, p):
                return JSONResponse({"detail": "Bu işlem için yetkiniz yok"}, status_code=403)
            return await call_next(request)
    if _kullanici_var():
        return JSONResponse({"detail": "yetkisiz", "giris": "parola"}, status_code=401)
    if API_TOKEN:
        return JSONResponse({"detail": "yetkisiz", "giris": "token"}, status_code=401)
    return await call_next(request)   # auth kapalı (yerel demo)


def _store():
    return open_store(cfg)


def _cameras() -> list[dict]:
    s = _store()
    try:
        return merged_cameras(cfg, s)
    finally:
        s.close()


def _sync_go2rtc() -> None:
    """Kameralardan go2rtc config'i üretir (canlı izleme fan-out'u).

    Kullanıcı go2rtc YAML'ı elle düzenlemez — kamera eklenince burada yenilenir.
    Dosya kaynakları ffmpeg: şemasıyla verilir (go2rtc döngüde oynatır).
    """
    rel = cfg.get("go2rtc.config_path", "")
    if not rel:
        return
    path = ROOT / rel
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = ["# Otomatik üretilir (src/server.py) — kamera eklendikçe yenilenir.",
             "streams:"]

    def _go2rtc_src(src: str, kamera_basliklari: str = "") -> str:
        if src.startswith(("http://", "https://")):
            # Kameraya özgü başlık varsa genel ayarı EZER — farklı sağlayıcılar
            # kendi Referer'ını şart koşar, yanlışı 403 döndürür
            basliklar = (kamera_basliklari
                         or cfg.get("stream.http_headers", "") or "").strip()
            if basliklar:
                # go2rtc'nin http istemcisi özel başlık göndermez → ffmpeg ile besle
                # -follow_redirects ffmpeg CLI'da yok (yalnız demuxer seçeneği); başlık
                # gönderildiğinde kaynak zaten yönlendirmiyor
                hdr = "\r\n".join(s.strip() for s in basliklar.splitlines() if s.strip())
                # -readrate 1.05: HLS'te ffmpeg indirdiği segmenti tek seferde
                # basar — röle 6 sn kare, 5.5 sn sessizlik üretiyordu (ölçüm:
                # 93 sn'de 15 donma, her biri ~5.5 sn; izleyicide "donuyor
                # sonra hızlanıyor"). Okuma gerçek zamana sabitlenir; 1.05,
                # kaynağa yetişememe birikimini önleyen küçük pay.
                return (f'exec:ffmpeg -readrate 1.05 -headers "{hdr}" -i "{src}"'
                        " -an -c copy -f rtsp {output}")
            # Başlıksız HTTP/HLS de aynı patlama sorununu yaşar → aynı tempo
            return (f'exec:ffmpeg -readrate 1.05 -i "{src}"'
                    " -an -c copy -f rtsp {output}")
        if src.startswith(("rtsp://", "rtmp://")):
            return src
        # Dosya kaynağı → sonsuz döngülü RTSP (gerçek kamera simülasyonu).
        # compose ./data/videos'u konteynerde /data/videos'a mount eder.
        # -an şart: sesli dosyada ffmpeg 8'in -re temposu bozuluyor (~0.56x
        # besleme → oynatıcı geride kalır, tampon boşalır, akış baştan başlar)
        return (f"exec:ffmpeg -re -stream_loop -1 -i /{src.lstrip('/')}"
                " -an -c copy -f rtsp {output}")

    for c in _cameras():
        # Değer JSON ile alıntılanır: içinde ": " geçen kaynak (ör. -headers "Referer: ...")
        # alıntısız yazılınca go2rtc'nin YAML ayrıştırıcısı TÜM config'i reddediyor.
        _hdr = str(c.get("http_headers") or "")
        lines.append(f"  {c['id']}: {json.dumps(_go2rtc_src(str(c['source']), _hdr))}")
        # Kamera duvarı substream'i: IP kameraların düşük çözünürlüklü ikinci akışı.
        # Duvarda 100 kareyi tam çözünürlükte çözmek tarayıcıyı boğar; tam çözünürlük
        # yalnız tek-kamera görünümü ve analiz içindir.
        if c.get("url_sub"):
            lines.append(f"  {c['id']}{SUB_SUFFIX}: {json.dumps(_go2rtc_src(str(c['url_sub']), _hdr))}")
    yeni = "\n".join(lines) + "\n"
    if path.exists() and path.read_text(encoding="utf-8") == yeni:
        return   # değişiklik yok → çalışan akışları kesme
    path.write_text(yeni, encoding="utf-8")

    # go2rtc çalışırken YAML'ı KENDİLİĞİNDEN okumaz; yeniden yüklemesi söylenmezse
    # "kamera ekledim ama canlıda görünmüyor" olur. Yalnız config değiştiğinde
    # tetiklenir (akışlar ~2 sn kesilir). exec: kaynakları çalışma-zamanı stream
    # API'sinden eklenemediği için (go2rtc güvenlik kısıtı) restart tek yol.
    import urllib.request

    base = (cfg.get("go2rtc.url", "") or "").rstrip("/")
    if not base:
        return
    try:
        urllib.request.urlopen(
            urllib.request.Request(f"{base}/api/restart", method="POST"), timeout=4).close()
    except Exception as e:
        print(f"[go2rtc] yeniden yükleme başarısız ({e}) — yeni kamera canlıda "
              f"görünmeyebilir, konteyneri yeniden başlatın", flush=True)


def _camera(camera_id: str) -> dict | None:
    return next((c for c in _cameras() if c.get("id") == camera_id), None)


# ─────────────────────────── Kimlik / RBAC ───────────────────────────

class GirisPayload(BaseModel):
    ad: str
    parola: str


class KullaniciPayload(BaseModel):
    ad: str
    parola: str
    rol: str = "izleyici"


def _istek_kullanicisi(request: Request) -> dict | None:
    return getattr(request.state, "kullanici", None)


def _yonetici_gerekli(request: Request) -> None:
    """Çerezle gelen ama yonetici olmayan isteği reddeder.

    Kimliksiz istek buraya yalnız auth tamamen kapalıyken düşer (middleware
    aksi hâlde 401 döndürür) — o kipte engel yok (yerel demo).
    """
    k = _istek_kullanicisi(request)
    if k and k.get("rol") != "yonetici":
        raise HTTPException(403, "Bu işlem yönetici yetkisi ister")


@app.post("/api/giris")
def api_giris(p: GirisPayload):
    s = _store()
    try:
        k = s.kullanici_bul(p.ad.strip())
    finally:
        s.close()
    if not k or not kimlik.parola_dogru(p.parola, k["parola_hash"]):
        # Ad mı parola mı yanlış söylenmez (kullanıcı adı taraması yemesin)
        raise HTTPException(401, "Kullanıcı adı veya parola hatalı")
    resp = JSONResponse({"ok": True, "ad": k["ad"], "rol": k["rol"]})
    resp.set_cookie(kimlik.OTURUM_CEREZ, kimlik.imzala(k["ad"], k["rol"]),
                    max_age=kimlik.OTURUM_SANIYE, httponly=True, samesite="lax")
    return resp


@app.post("/api/cikis")
def api_cikis():
    resp = JSONResponse({"ok": True})
    resp.delete_cookie(kimlik.OTURUM_CEREZ)
    return resp


@app.get("/api/ben")
def api_ben(request: Request):
    """UI'nin kim olduğunu ve giriş kipini öğrendiği uç."""
    k = _istek_kullanicisi(request) or {}
    return {"ad": k.get("ad"), "rol": k.get("rol"),
            "kullanici_tanimli": _kullanici_var()}


@app.get("/api/kullanicilar")
def api_kullanicilar(request: Request):
    _yonetici_gerekli(request)
    s = _store()
    try:
        return s.kullanici_listele()
    finally:
        s.close()


@app.post("/api/kullanicilar")
def api_kullanici_ekle(p: KullaniciPayload, request: Request):
    _yonetici_gerekli(request)
    ad = p.ad.strip()
    if not ad or len(ad) > 40:
        raise HTTPException(400, "Kullanıcı adı 1-40 karakter olmalı")
    if len(p.parola) < 6:
        raise HTTPException(400, "Parola en az 6 karakter olmalı")
    rol = p.rol if p.rol in kimlik.ROLLER else "izleyici"
    s = _store()
    try:
        if s.kullanici_sayisi() == 0:
            rol = "yonetici"   # ilk kullanıcı kilidi: sistem yöneticisiz kalamaz
        s.kullanici_ekle(ad, kimlik.parola_hashle(p.parola), rol)
    finally:
        s.close()
    _kullanici_cache_sifirla()
    return {"ok": True, "ad": ad, "rol": rol}


@app.delete("/api/kullanicilar/{ad}")
def api_kullanici_sil(ad: str, request: Request):
    _yonetici_gerekli(request)
    s = _store()
    try:
        hedef = s.kullanici_bul(ad)
        if not hedef:
            raise HTTPException(404, "Kullanıcı bulunamadı")
        if hedef["rol"] == "yonetici":
            kalan = sum(1 for u in s.kullanici_listele() if u["rol"] == "yonetici")
            if kalan <= 1:
                raise HTTPException(400, "Son yönetici silinemez — önce başka yönetici ekleyin")
        s.kullanici_sil(ad)
    finally:
        s.close()
    _kullanici_cache_sifirla()
    return {"ok": True}


@app.get("/api/cameras")
def api_cameras():
    return _cameras()


class CameraPayload(BaseModel):
    name: str
    source: str
    id: str = ""
    source_sub: str = ""   # kamera duvarı için düşük çözünürlüklü akış (opsiyonel)
    http_headers: str = ""  # bazı HLS/CDN sağlayıcıları kendi Referer'ını şart koşar


def _slug(s: str) -> str:
    import re
    tr = str.maketrans("çğıiöşüÇĞIİÖŞÜ", "cgiiosucgiiosu")
    return re.sub(r"[^a-z0-9]+", "-", s.translate(tr).lower()).strip("-") or "kamera"


@app.post("/api/cameras")
def api_add_camera(p: CameraPayload):
    s = _store()
    try:
        cid = p.id or _slug(p.name)
        base = cid; i = 2
        existing = {c["id"] for c in _cameras()}
        while cid in existing:
            cid = f"{base}-{i}"; i += 1
        s.add_camera(cid, p.name, p.source, p.source_sub.strip(),
                     p.http_headers.strip())
        akis.kaydet(p.source, p.http_headers.strip())
        _sync_go2rtc()
        return {"ok": True, "id": cid}
    finally:
        s.close()


class CameraPatch(BaseModel):
    name: str | None = None
    source: str | None = None
    source_sub: str | None = None
    http_headers: str | None = None


@app.patch("/api/cameras/{cid}")
def api_update_camera(cid: str, p: CameraPatch):
    """Var olan kamerayı günceller (ad/kaynak/substream).

    POST /api/cameras yalnız EKLER (aynı id gelirse yeni id türetir); kurulumdan
    sonra adres/substream değiştirmek için düzenleme ucu gerekir.
    """
    cam = _camera(cid)
    if not cam:
        raise HTTPException(404, "Kamera bulunamadı")
    s = _store()
    try:
        s.add_camera(cid,
                     (p.name or cam["name"]).strip(),
                     (p.source if p.source is not None else cam["source"]).strip(),
                     (p.source_sub if p.source_sub is not None
                      else (cam.get("url_sub") or "")).strip(),
                     # add_camera upsert'tir: geçilmeyen alan NULL'a düşer.
                     # Başlığı taşımazsak düzenleme onu SİLER.
                     (p.http_headers if p.http_headers is not None
                      else (cam.get("http_headers") or "")).strip())
        # Görevlere dokunulmaz: add_camera'nın upsert'i tasks sütununu yazmaz.
        _sync_go2rtc()
        return {"ok": True, "id": cid}
    finally:
        s.close()


def _safe_src(s: str) -> str:
    """Hata metninden kimlik bilgisini siler (log/UI'a şifre sızmasın)."""
    import re
    return re.sub(r"//[^/@\s]+:[^/@\s]+@", "//***:***@", s or "")


class ProbePayload(BaseModel):
    source: str


@app.post("/api/cameras/probe")
def api_probe_camera(p: ProbePayload):
    """Kamerayı KAYDETMEDEN dener: erişilebilir mi, çözünürlük/fps/codec ne, örnek kare.

    Sahada en sık kaybedilen zaman "ekledim ama görüntü yok" turudur; doğrulama
    ekleme anında yapılır. Kimlik bilgisi yanıtta ve logda maskelenir.
    """
    import base64

    src = (p.source or "").strip()
    if not src:
        raise HTTPException(400, "Kaynak boş")
    if not src.startswith(("rtsp://", "rtmp://", "http://", "https://")):
        f = (ROOT / src).resolve()
        if not f.is_file():
            return {"ok": False, "error": f"Dosya bulunamadı: {src}"}
        src = str(f)

    import av

    # stimeout: RTSP soketi yanıt vermezse sonsuza kadar asılı kalmasın (µs)
    opts = {"rtsp_transport": "tcp", "stimeout": "6000000"} if src.startswith("rtsp") else {}
    opts.update(http_options(cfg, src))
    # HLS/HTTP çok daha yavaş açılır: master playlist → varyant → ilk segmentler
    # (ölçüldü: kamu test akışında 43 sn). RTSP/dosya için kısa süre yeterli.
    sure = 60 if src.startswith(("http://", "https://")) else 12
    try:
        with av.open(src, options=opts, timeout=sure) as c:
            if not c.streams.video:
                return {"ok": False, "error": "Kaynakta video akışı yok"}
            vs = c.streams.video[0]
            frame = next((f for f in c.decode(vs)), None)
            if frame is None:
                return {"ok": False, "error": "Bağlanıldı ama kare çözülemedi (codec/anahtar kare?)"}
            img = frame.to_ndarray(format="bgr24")
            h, w = img.shape[:2]
            if w > 640:
                img = cv2.resize(img, (640, int(h * 640 / w)))
            ok, buf = cv2.imencode(".jpg", img, [int(cv2.IMWRITE_JPEG_QUALITY), 82])
            return {"ok": True, "width": w, "height": h,
                    "fps": round(float(vs.average_rate or 0), 1),
                    "codec": (vs.codec_context.name or "").upper(),
                    "preview": base64.b64encode(buf).decode() if ok else ""}
    except Exception as e:
        msg = _safe_src(str(e)) or e.__class__.__name__
        low = msg.lower()
        if "401" in msg or "unauthorized" in low:
            msg = "Kimlik doğrulama reddedildi — kullanıcı adı/şifre hatalı"
        elif "404" in msg or "not found" in low:
            msg = "Akış yolu bulunamadı (404) — RTSP yolunu kamera markasına göre kontrol edin"
        elif "connection refused" in low or "no route" in low:
            msg = "Adrese ulaşılamadı — IP/port ve ağ erişimini kontrol edin"
        elif "immediate exit" in low or "timed out" in low or "timeout" in low:
            msg = "Zaman aşımı — kamera yanıt vermedi (IP doğru mu, cihaz ağda mı?)"
        return {"ok": False, "error": msg}


@app.delete("/api/cameras/{cid}")
def api_del_camera(cid: str):
    s = _store()
    try:
        s.delete_camera(cid)
        _sync_go2rtc()
        return {"ok": True}
    finally:
        s.close()


class TasksPayload(BaseModel):
    tasks: dict


@app.post("/api/cameras/{cid}/tasks")
def api_set_tasks(cid: str, p: TasksPayload):
    cam = _camera(cid)
    if not cam:
        raise HTTPException(404, "Kamera bulunamadı")
    # Gönderilmeyen anahtar mevcut değeri KORUR (yoksa varsayılan). Eski davranış
    # False'a düşürüyordu: DEFAULT_TASKS'a "record" eklenince, görev değiştiren
    # her tıklama kaydı sessizce kapatırdı.
    eski_tasks = cam.get("tasks") or {}
    tasks = {k: bool(p.tasks[k]) if k in p.tasks
             else bool(eski_tasks.get(k, DEFAULT_TASKS[k]))
             for k in DEFAULT_TASKS}
    s = _store()
    try:
        # config kamerası DB'de yoksa önce upsert (görevler DB'de yaşar).
        # url_sub ve http_headers da taşınır: upsert hepsini yazar, geçilmeyen
        # alan NULL'a düşüp SİLİNİRDİ (PATCH ucunda da aynı hata vardı).
        s.add_camera(cid, cam["name"], cam["source"], cam.get("url_sub") or "",
                     cam.get("http_headers") or "")
        s.set_camera_tasks(cid, tasks)
        return {"ok": True, "tasks": tasks}
    finally:
        s.close()


@app.get("/api/recordings")
def api_recordings(camera: str = "", start: str = "", end: str = "",
                   limit: int = Query(2000, ge=1, le=20000)):
    """Kayıt segmentleri — zaman çizelgesi ve geri oynatma buradan beslenir.

    Segment süreleri ffprobe ile ÖLÇÜLMÜŞTÜR; nominal süreye güvenen bir zaman
    çizelgesi anahtar-kare kesimleri yüzünden zamanla kayar.
    """
    s = _store()
    try:
        return s.list_recordings(camera, start or None, end or None, limit)
    finally:
        s.close()


_VOD_MAX = 7200   # sn — tek playlist penceresi tavanı (bkz. aşağıdaki not)


@app.get("/api/vod/{camera}/index.m3u8")
def api_vod(camera: str, start: str, end: str = "", token: str = ""):
    """Geçmişe dönük oynatma için dinamik HLS VOD playlist'i.

    Neden playlist: kayıt 60 sn'lik ayrı mp4 segmentlerinde; tarayıcı bunları tek
    sürekli video gibi oynatsın diye HLS listesi üretilir (hls.js).

    İki sert kural (sektör deneyimi):
      * Pencere ≤ 2 saat. Binlerce segmentlik tek playlist oynatıcıları çökertiyor.
      * Ardışık segmentler arasına EXT-X-DISCONTINUITY KONMAZ — aynı kameradan
        remux edilmiş segmentler uyumludur; discontinuity hls.js'te takılma yapar.
        Gerçek boşluk playlist'e hiç girmez (UI zaman çizelgesinde boşluk gösterir).
    """
    s = _store()
    try:
        segler = s.list_recordings(camera, start, end or None, limit=4000)
    finally:
        s.close()
    if not segler:
        raise HTTPException(404, "Bu aralıkta kayıt yok")

    tok = f"?token={token}" if token else ""
    satirlar = ["#EXTM3U", "#EXT-X-VERSION:7", "#EXT-X-PLAYLIST-TYPE:VOD"]
    govde: list[str] = []
    toplam = 0.0
    enbuyuk = 0.0
    for r in segler:
        sure = float(r["duration"] or 0)
        if sure <= 0:
            continue
        if toplam + sure > _VOD_MAX:
            break
        toplam += sure
        enbuyuk = max(enbuyuk, sure)
        # EXTINF ÖLÇÜLEN süredir; nominal yazılırsa zaman çizelgesi kayar
        govde.append(f"#EXTINF:{sure:.3f},")
        govde.append(f"/media/{cfg.get('record.dir', 'rec')}/{r['path']}{tok}")
    if not govde:
        raise HTTPException(404, "Oynatılabilir segment yok")
    satirlar.append(f"#EXT-X-TARGETDURATION:{int(enbuyuk) + 1}")
    satirlar.extend(govde)
    satirlar.append("#EXT-X-ENDLIST")
    return Response("\n".join(satirlar) + "\n",
                    media_type="application/vnd.apple.mpegurl",
                    headers={"Cache-Control": "no-store"})


@app.get("/api/recordings/timeline")
def api_timeline(camera: str, start: str, end: str):
    """Zaman çizelgesi bandı: kayıt olan/olmayan aralıklar (video indirmeden).

    Ardışık segmentler birleştirilir; aradaki kopukluk boşluk olarak döner —
    operatör "burada kayıt yok" bilgisini görmeli, sessizce atlanmamalı.
    """
    s = _store()
    try:
        segler = s.list_recordings(camera, start, end, limit=20000)
    finally:
        s.close()
    bloklar: list[dict] = []
    for r in segler:
        b, e = r["start_time"], r["end_time"]
        if bloklar and _yakin(bloklar[-1]["end"], b):
            bloklar[-1]["end"] = e
            bloklar[-1]["segments"] += 1
        else:
            bloklar.append({"start": b, "end": e, "segments": 1})
    return {"camera": camera, "blocks": bloklar}


def _yakin(a: str, b: str, tolerans: float = 5.0) -> bool:
    """İki zaman damgası bitişik sayılır mı (küçük remux boşluklarını yut)."""
    try:
        da = datetime.fromisoformat(str(a).replace(" ", "T"))
        db = datetime.fromisoformat(str(b).replace(" ", "T"))
        return abs((db - da).total_seconds()) <= tolerans
    except ValueError:
        return False


@app.get("/api/discover")
def api_discover(timeout: float = Query(4.0, ge=1.0, le=15.0)):
    """Yerel ağdaki ONVIF kameralarını bulur (WS-Discovery multicast).

    Yalnız yerel ağ taranır; internete istek gitmez. ONVIF'i kapalı kameralar
    görünmez — onlar için marka şablonu/yol tarama kullanılır.
    """
    from .discovery import onvif_kesfet

    try:
        bulunan = onvif_kesfet(timeout)
    except Exception as e:
        raise HTTPException(500, f"Keşif başarısız: {e}")
    mevcut = {str(c.get("source", "")) for c in _cameras()}
    for b in bulunan:
        b["ekli"] = any(b["ip"] in s for s in mevcut)
    return {"devices": bulunan, "count": len(bulunan)}


class OnvifPayload(BaseModel):
    ip: str
    username: str = ""
    password: str = ""
    port: int = 80


@app.post("/api/discover/onvif")
def api_onvif(p: OnvifPayload):
    """Kameranın KENDİ bildirdiği akış adreslerini alır (ana + substream)."""
    from .discovery import onvif_akislari

    r = onvif_akislari(p.ip.strip(), p.username, p.password, p.port)
    if not r.get("ok"):
        return {"ok": False, "error": _safe_src(r.get("hata", "bilinmeyen"))}
    return r


class ProbePathsPayload(BaseModel):
    ip: str
    username: str = ""
    password: str = ""
    port: int = 554
    brand: str = ""


@app.post("/api/discover/paths")
def api_probe_paths(p: ProbePathsPayload):
    """ONVIF yoksa: bilinen RTSP yollarını dener, GERÇEKTEN AÇILANLARI döndürür."""
    from .discovery import MARKA_YOLLARI, yol_dene

    if p.brand and p.brand not in MARKA_YOLLARI:
        raise HTTPException(422, f"Bilinmeyen marka: {p.brand}")
    try:
        bulunan = yol_dene(p.ip.strip(), p.username, p.password, p.port, p.brand)
    except Exception as e:
        raise HTTPException(500, f"Tarama başarısız: {_safe_src(str(e))}")
    return {"found": [{**b, "url": _safe_src(b["url"]), "_url": b["url"]} for b in bulunan],
            "count": len(bulunan)}


@app.get("/api/discover/brands")
def api_brands():
    from .discovery import MARKA_YOLLARI
    return {"brands": sorted(MARKA_YOLLARI)}


class ExportPayload(BaseModel):
    camera: str
    start: str
    end: str


@app.post("/api/export")
def api_export(p: ExportPayload, request: Request):
    """Zaman aralığını tek mp4 + SHA-256 imzalı manifest olarak dışa aktarır.

    Kanıt zinciri: manifest hem birleşik dosyanın hem kaynak segmentlerin
    özetini taşır — dosya sonradan değişirse özet tutmaz.
    """
    from .exporter import disa_aktar

    s = _store()
    try:
        segler = s.list_recordings(p.camera, p.start, p.end, limit=20000)
    finally:
        s.close()
    try:
        # Kanıt manifestindeki "isteyen" artık gerçek kullanıcı (denetim izi)
        kim = (_istek_kullanicisi(request) or {}).get("ad") or "operatör"
        return disa_aktar(cfg, p.camera, segler, isteyen=kim)
    except ValueError as e:
        raise HTTPException(404, str(e))
    except Exception as e:
        raise HTTPException(500, f"Dışa aktarma başarısız: {e}")


@app.get("/api/recordings/stats")
def api_recordings_stats():
    """Kamera başına arşiv derinliği ve disk tüketimi."""
    s = _store()
    try:
        rows = s.recordings_stats()
        for r in rows:
            r["oldest"] = str(r["oldest"]) if r.get("oldest") else None
            r["newest"] = str(r["newest"]) if r.get("newest") else None
            r["bytes"] = int(r.get("bytes") or 0)
        return {"cameras": rows, "total_bytes": s.recordings_size(),
                "keep_days": int(cfg.get("record.keep_days", 15)),
                "enabled": bool(cfg.get("record.enabled", True))}
    finally:
        s.close()


@app.get("/api/rapor")
def api_rapor(start: str = Query(...), end: str = Query(...), camera: str = ""):
    """Kamera bazlı günlük faaliyet raporu (TRASSIR sınıfı raporlama).

    Satır = kamera × gün: giriş/çıkış, plaka okuma + benzersiz plaka, yüz,
    alarm, kayıt kapsaması (dk). Tarihler yerel gün olarak yorumlanır — UI
    ISO (UTC) çevirip gönderir; kayıtlardaki gün kayması dersinin aynısı.
    """
    s = _store()
    try:
        def _g(sql, *p):
            try:
                return s._all(sql, p)
            except Exception:
                return []
        oz: dict = {}
        def _al(anahtar):
            def ekle(rows, ad):
                for r in rows:
                    k = (str(r["gun"])[:10], r["camera_id"])
                    oz.setdefault(k, {})[ad] = r["n"]
            return ekle
        kos = " AND camera_id=?" if camera else ""
        ek = (camera,) if camera else ()
        ekle = _al(oz)
        ekle(_g(f"SELECT date(time) gun, camera_id, count(*) n FROM count_events"
                f" WHERE time>=? AND time<=? AND direction='in'{kos}"
                f" GROUP BY 1,2", start, end, *ek), "giris")
        ekle(_g(f"SELECT date(time) gun, camera_id, count(*) n FROM count_events"
                f" WHERE time>=? AND time<=? AND direction='out'{kos}"
                f" GROUP BY 1,2", start, end, *ek), "cikis")
        ekle(_g(f"SELECT date(time) gun, camera_id, count(*) n FROM plate_events"
                f" WHERE time>=? AND time<=?{kos} GROUP BY 1,2", start, end, *ek), "plaka")
        ekle(_g(f"SELECT date(time) gun, camera_id, count(DISTINCT plate) n FROM plate_events"
                f" WHERE time>=? AND time<=?{kos} GROUP BY 1,2", start, end, *ek), "benzersiz_plaka")
        ekle(_g(f"SELECT date(time) gun, camera_id, count(*) n FROM face_events"
                f" WHERE time>=? AND time<=?{kos} GROUP BY 1,2", start, end, *ek), "yuz")
        ekle(_g(f"SELECT date(time) gun, camera_id, count(*) n FROM alerts"
                f" WHERE time>=? AND time<=?{kos} GROUP BY 1,2", start, end, *ek), "alarm")
        ekle(_g(f"SELECT date(start_time) gun, camera_id,"
                f" CAST(sum(duration)/60 AS INTEGER) n FROM recordings"
                f" WHERE start_time>=? AND start_time<=?{kos} GROUP BY 1,2",
                start, end, *ek), "kayit_dk")
        satirlar = [{"gun": g, "camera_id": c, **v} for (g, c), v in sorted(oz.items())]
        return {"satirlar": satirlar}
    finally:
        s.close()


@app.get("/api/rapor.xlsx")
def api_rapor_xlsx(start: str = Query(...), end: str = Query(...), camera: str = ""):
    """Aynı rapor, Excel dosyası olarak (openpyxl — saf Python, CDN'siz)."""
    veri = api_rapor(start, end, camera)["satirlar"]
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook(); ws = wb.active; ws.title = "Faaliyet"
    basliklar = ["Gün", "Kamera", "Giriş", "Çıkış", "Plaka", "Benzersiz plaka",
                 "Yüz", "Alarm", "Kayıt (dk)"]
    ws.append(basliklar)
    for h in ws[1]: h.font = Font(bold=True)
    adlar = {c["id"]: c["name"] for c in _cameras()}
    for r in veri:
        ws.append([r["gun"], adlar.get(r["camera_id"], r["camera_id"]),
                   r.get("giris", 0), r.get("cikis", 0), r.get("plaka", 0),
                   r.get("benzersiz_plaka", 0), r.get("yuz", 0),
                   r.get("alarm", 0), r.get("kayit_dk", 0)])
        for i, b in enumerate(basliklar, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = max(11, len(b) + 3)
    buf = io.BytesIO(); wb.save(buf)
    return Response(buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="aurasvision-rapor-{start[:10]}-{end[:10]}.xlsx"'})


class SearchPayload(BaseModel):
    q: str
    camera: str = ""
    limit: int = 24


@app.post("/api/search")
def api_search(p: SearchPayload):
    """Görünüm araması: serbest metin → arşivdeki nesne kırpmaları.

    Vektörler worker'da üretilir (gpu_engine._run_vektor); burada yalnız metin
    kodlanır ve pgvector/SQLite'ta en yakınlar bulunur. Model ilk çağrıda
    yüklenir (~2 sn), sonrası milisaniyeler.
    """
    q = p.q.strip()
    if len(q) < 2:
        raise HTTPException(422, "Arama metni çok kısa")
    from . import arama
    try:
        vec = arama.metin_vektoru(q)
    except Exception as e:
        raise HTTPException(503, f"Arama modeli yüklenemedi: {e}")
    s = _store()
    try:
        rows = s.search_nesne_vektor(vec, limit=min(max(p.limit, 1), 60),
                                     camera_id=p.camera)
    finally:
        s.close()
    return {"sorgu": q, "sonuc": rows}


@app.get("/api/recordings/root")
def api_recordings_root():
    """Arşiv kök klasörünün tam yolu — operatör dosyalara elle erişebilsin."""
    from .recorder import kayit_kok
    return {"path": str(kayit_kok(cfg))}


@app.post("/api/recordings/open-folder")
def api_recordings_open(camera: str = "", gun: str = ""):
    """Arşiv klasörünü SUNUCU makinesinin dosya yöneticisinde açar.

    Tek makine kurulumunda operatör panelin koştuğu bilgisayardadır — "dosyaya
    nasıl erişirim" sorusunun doğrudan cevabı. Uzak erişimde pencere sunucuda
    açılır; UI bu yüzden tam yolu da gösterir (kopyalanabilir).
    """
    import subprocess, sys
    from .recorder import kayit_kok
    kok = kayit_kok(cfg)
    hedef = kok
    if camera:
        aday = kok / camera if not gun else kok / camera / gun
        if aday.is_dir():
            hedef = aday
    if not hedef.is_dir():
        raise HTTPException(404, "Arşiv klasörü henüz yok")
    try:
        if sys.platform == "win32":
            subprocess.Popen(["explorer", str(hedef)])
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(hedef)])
        else:
            subprocess.Popen(["xdg-open", str(hedef)],
                             stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        return {"ok": True, "path": str(hedef)}
    except FileNotFoundError:
        # masaüstü ortamı yok (headless sunucu) — yol yine döner, UI gösterir
        return {"ok": False, "path": str(hedef),
                "detail": "Dosya yöneticisi bulunamadı — yolu kopyalayıp elle açın"}


@app.get("/api/status")
def api_status():
    """Sistem bileşenlerinin GERÇEK durumu (sidebar göstergesi buradan beslenir).

    Kenar çubuğundaki gösterge eskiden sabit "aktif" yazıyordu — sistem çökse bile
    yeşil görünürdü. Operatör paneli yalan söylememeli.
    """
    import time as _t

    bilesenler = []

    def olc(ad: str, fn) -> None:
        t0 = _t.monotonic()
        try:
            detay = fn() or ""
            ok = True
        except Exception as e:
            detay, ok = str(e)[:120], False
        bilesenler.append({"ad": ad, "ok": ok, "detay": detay,
                           "ms": round((_t.monotonic() - t0) * 1000)})

    def _db():
        s = _store()
        try:
            s.latest_health()
            return "PostgreSQL" if os.environ.get("DATABASE_URL") or cfg.get("db.url", "") else "SQLite"
        finally:
            s.close()

    def _redis():
        from .bus import open_bus
        r = open_bus(cfg)
        if r is None:
            # Eksiklik değil, kurulum profili: tek makinede analiz servisi
            # olayları doğrudan veritabanına yazar (ayrı ingestor yoktur).
            return "tek makine — doğrudan yazım"
        r.ping()
        return "Redis bağlı"

    def _go2rtc():
        import urllib.request
        base = (cfg.get("go2rtc.url", "") or "").rstrip("/")
        if not base:
            return "yapılandırılmadı"
        with urllib.request.urlopen(f"{base}/api/streams", timeout=3) as r:
            return f"{len(json.loads(r.read()))} akış"

    def _ingestor():
        """Olay işleyici birikmesi — kuyruk büyüyorsa olaylar DB'ye GEÇMİYOR demektir.

        Sahada yaşandı: ingestor ölünce 11k+ olay stream'de birikti, panel
        "sağlıklı" gösterdi. Artık consumer group'un gecikmesi ölçülür.
        """
        from .bus import GROUP, STREAM, open_bus
        r = open_bus(cfg)
        if r is None:
            return "tek makine — ayrı işleyici yok"
        gruplar = {g.get("name"): g for g in r.xinfo_groups(STREAM)}
        g = gruplar.get(GROUP)
        if g is None:
            raise RuntimeError("olay işleyici hiç bağlanmamış (consumer group yok)")
        bekleyen = int(g.get("pending") or 0)
        lag = g.get("lag")   # Redis 7+: hiç teslim edilmemiş mesaj sayısı
        birikme = bekleyen + int(lag or 0)
        if birikme > 1000:
            raise RuntimeError(f"{birikme} olay kuyrukta birikti — ingestor yetişmiyor/ölü")
        return "birikme yok" if birikme == 0 else f"{birikme} olay işleniyor"

    def _disk():
        """Arşiv diski — dolarsa kayıt DA analiz DE sessizce durur."""
        import shutil
        kok = ROOT / cfg.get("paths.output_dir", "output") / cfg.get("record.dir", "rec")
        d = shutil.disk_usage(kok if kok.exists() else ROOT)
        bos_gb = d.free / (1024 ** 3)
        detay = f"boş {bos_gb:.0f} GB / toplam {d.total / (1024 ** 3):.0f} GB"
        if bos_gb < 10 or d.free / d.total < 0.05:
            raise RuntimeError(f"disk doluyor — {detay}")
        return detay

    def _arama():
        if not cfg.get("arama.enabled", True):
            return "kapalı"
        s = _store()
        try:
            rows = s._all("SELECT COUNT(*) AS n, MAX(time) AS son FROM nesne_vektor")
        finally:
            s.close()
        n, son = int(rows[0]["n"] or 0), rows[0].get("son")
        if not n:
            return "henüz vektör yok (nesne görülünce başlar)"
        return f"{n:,} nesne dizinde · son yazım {str(son)[:19]}".replace(",", ".")

    olc("Veritabanı", _db)
    olc("Olay yolu", _redis)
    olc("Olay işleyici", _ingestor)
    olc("Canlı akış", _go2rtc)
    olc("Disk alanı", _disk)
    olc("Arama dizini", _arama)

    # Worker sağlığı: heartbeat TAZE olması yetmez — kare de ÜRETİYOR olmalı.
    # Ölü bir kamerayı işlemeye çalışan worker canlı heartbeat atar ama 0 fps
    # üretir; eski sürümde bu "sağlıklı" görünüyordu. Sahada en pahalı yanılgı bu.
    s = _store()
    try:
        simdi = datetime.now(timezone.utc)
        taze, ureten, durgun = 0, 0, []
        for h in s.latest_health():
            t = str(h.get("time") or "")
            try:
                d = datetime.fromisoformat(t.replace(" ", "T"))
                if d.tzinfo is None:
                    d = d.replace(tzinfo=timezone.utc)
                if (simdi - d).total_seconds() >= 15:
                    continue
            except ValueError:
                continue
            taze += 1
            if float(h.get("fps") or 0) > 0.1:
                ureten += 1
            else:
                durgun.append(str(h.get("camera_id") or "?"))
        kamera = len(_cameras())
    finally:
        s.close()

    if not taze:
        detay, ok = "çalışmıyor — sürekli analiz yok", False
    elif ureten == 0:
        detay, ok = (f"{taze} kamera bağlı ama KARE ÜRETMİYOR "
                     f"({', '.join(durgun[:3])}) — kaynak erişilemiyor olabilir"), False
    elif durgun:
        detay, ok = (f"{ureten}/{taze} kamera işleniyor · durgun: "
                     f"{', '.join(durgun[:3])}"), False
    else:
        detay, ok = f"{ureten}/{kamera} kamera işleniyor", True
    bilesenler.append({"ad": "Analiz worker", "ok": ok, "detay": detay, "ms": 0})

    # Kayıt servisi ayrı bileşen: mevzuat gereği çalışıyor olmalı, sessizce
    # durması yükümlülüğün karşılanmaması demektir.
    if cfg.get("record.enabled", True):
        s = _store()
        try:
            son = s.recordings_stats()
        finally:
            s.close()
        yeni = 0
        for r in son:
            try:
                d = datetime.fromisoformat(str(r.get("newest") or "").replace(" ", "T"))
                if d.tzinfo is None:
                    d = d.replace(tzinfo=timezone.utc)
                if (datetime.now(timezone.utc) - d).total_seconds() < 300:
                    yeni += 1
            except ValueError:
                continue
        toplam_gb = sum(int(r.get("bytes") or 0) for r in son) / (1024 ** 3)
        bilesenler.append({
            "ad": "Kayıt servisi", "ok": yeni > 0,
            "detay": (f"{yeni} kamera kaydediliyor · arşiv {toplam_gb:.1f} GB"
                      if yeni else "son 5 dakikada yeni kayıt YOK"),
            "ms": 0})

    return {"ok": all(b["ok"] for b in bilesenler), "bilesenler": bilesenler}


@app.get("/api/health")
def api_health():
    s = _store()
    try:
        return s.latest_health()
    finally:
        s.close()


@app.get("/api/sysinfo")
def api_sysinfo(request: Request):
    go2rtc = cfg.get("go2rtc.url", "")
    if go2rtc.startswith(("http://localhost", "http://127.0.0.1")):
        go2rtc = f"{request.url.scheme}://{request.url.hostname}:1984"
    return {"go2rtc": go2rtc}


@app.websocket("/api/stream")
async def api_stream(ws: WebSocket):
    """Canlı akışı go2rtc'den vekiller (MSE/WebRTC sinyalleşmesi dahil).

    Neden vekil: (1) go2rtc çapraz-origin WebSocket'i 403'le reddeder, (2) tarayıcıyı
    doğrudan 1984'e bağlamak go2rtc'nin KİMLİK DOĞRULAMASIZ API'sini her operatörün
    ağına açardı — kritik sahada kabul edilemez. Burada akış AURAS_TOKEN'a bağlanır
    ve müşteri yalnız tek portu (uygulama) dışarı açar.
    """
    import asyncio
    from urllib.parse import quote

    q = ws.query_params
    # HTTP middleware'i WebSocket kapsamına uygulanmaz → yetki burada kontrol edilir.
    # Kabul sırası middleware ile aynı: makine token'ı → oturum çerezi → auth kapalı.
    oturum = kimlik.coz(ws.cookies.get(kimlik.OTURUM_CEREZ, ""))
    izinli = (API_TOKEN and secrets.compare_digest(q.get("token", ""), API_TOKEN)) \
        or (oturum is not None and oturum["ad"] in _kullanici_rolleri()) \
        or (not API_TOKEN and not _kullanici_var())
    if not izinli:
        await ws.close(code=1008)   # policy violation
        return
    src = q.get("src", "")
    base = (cfg.get("go2rtc.url", "") or "").rstrip("/")
    if not src or not base:
        await ws.close(code=1011)
        return

    import websockets

    upstream = "ws" + base[4:] + f"/api/ws?src={quote(src)}"
    await ws.accept()
    up = None
    try:
        # max_size=None: fMP4 segmentleri (1440p anahtar kare) 1MB varsayılanını aşabilir
        up = await websockets.connect(upstream, max_size=None)

        async def to_upstream() -> None:
            while True:                      # istemci yalnız JSON kontrol mesajı yollar
                await up.send(await ws.receive_text())

        async def to_client() -> None:
            async for m in up:               # ikili = video segmenti, metin = kontrol
                await (ws.send_bytes(m) if isinstance(m, bytes) else ws.send_text(m))

        tasks = [asyncio.create_task(to_upstream()), asyncio.create_task(to_client())]
        done, pending = await asyncio.wait(tasks, return_when=asyncio.FIRST_COMPLETED)
        for t in pending:
            t.cancel()
        # İptalin TAMAMLANMASINI bekle: okuyucu görev hâlâ askıdayken close() çağrılırsa
        # el sıkışma tamamlanmaz ve go2rtc tarafında ölü tüketici kalır (kamera başına sızıntı).
        await asyncio.gather(*pending, return_exceptions=True)
        for t in done:
            t.exception()   # istisnayı tüket (uyarı basılmasın); kopma normaldir
    except WebSocketDisconnect:
        pass
    except Exception as e:
        print(f"[stream] {src} vekil hatası: {e}", flush=True)
    finally:
        # Yukarı akış HER durumda kapanır — istemci sekmeyi aniden kapatsa bile
        if up is not None:
            try:
                await up.close()
            except Exception:
                pass
        try:
            await ws.close()
        except RuntimeError:
            pass   # zaten kapanmış


# go2rtc oynatıcı bileşeni: CORS başlığı göndermediği için tarayıcı onu başka
# origin'den ES module olarak YÜKLEYEMEZ → kendi origin'imizden vekilleriz.
# Repoya kopyalamak yerine vekil: bileşen daima çalışan go2rtc sürümüyle eşleşir.
_VENDOR_CACHE: dict[str, bytes] = {}
_VENDOR_ALLOWED = {"video-rtc.js"}


@app.get("/vendor/{name}")
def vendor_asset(name: str):
    if name not in _VENDOR_ALLOWED:
        raise HTTPException(404, "bilinmeyen kaynak")
    if name not in _VENDOR_CACHE:
        import urllib.request

        base = (cfg.get("go2rtc.url", "") or "").rstrip("/")
        if not base:
            raise HTTPException(503, "go2rtc yapılandırılmadı")
        try:
            with urllib.request.urlopen(f"{base}/{name}", timeout=5) as r:
                _VENDOR_CACHE[name] = r.read()
        except Exception as e:
            raise HTTPException(502, f"go2rtc'den alınamadı: {e}")
    return Response(_VENDOR_CACHE[name], media_type="text/javascript",
                    headers={"Cache-Control": "max-age=3600"})


# Snapshot TTL cache — VideoCapture pahalı; aynı kareyi N sn tekrar üretme (perf).
_SNAP_CACHE: dict[str, tuple[float, bytes]] = {}
_SNAP_TTL = 12.0  # saniye — UI yenileme aralığından KISA olmamalı, yoksa her istek
                  # önbelleği ıskalayıp video açar ve iş parçacığı havuzu dolar


def _grab_jpeg(source: str) -> bytes | None:
    # Zaman aşımı ŞART: erişilemeyen bir ağ kamerası (kopmuş RTSP, süresi dolmuş
    # HLS) varsayılan ayarla 30 sn boyunca iş parçacığını tutuyor. Çok kameralı
    # ekranda birkaç ölü kamera FastAPI'nin havuzunu doldurup TÜM API'yi
    # kilitledi (mobil arayüzde yaşandı). Ölü kamera hızlı başarısız olmalı.
    cap = akis.ac(source, cfg, timeout_ms=5000)
    if not cap.isOpened():
        return None
    total = int(cap.get(cv2.CAP_PROP_FRAME_COUNT) or 0)
    if total > 1:
        cap.set(cv2.CAP_PROP_POS_FRAMES, total // 2)
    ok, frame = cap.read()
    cap.release()
    if not ok:
        return None
    ok, buf = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 82])
    return buf.tobytes() if ok else None


@app.get("/api/snapshot")
def api_snapshot(camera: str = Query(...), fresh: int = 0):
    cam = _camera(camera)
    if not cam:
        raise HTTPException(404, "Kamera bulunamadı")
    now = time.time()
    cached = _SNAP_CACHE.get(camera)
    if not fresh and cached and now - cached[0] < _SNAP_TTL:
        data = cached[1]
    else:
        data = _grab_jpeg(cam["source"])
        if data is None:
            raise HTTPException(400, "Kare okunamadı")
        _SNAP_CACHE[camera] = (now, data)
    return Response(content=data, media_type="image/jpeg",
                    headers={"Cache-Control": "public, max-age=4"})


class ZonePayload(BaseModel):
    camera: str
    zones: list[dict]   # [{kind, name, points:[[x,y]..], classes:[..], direction}]


@app.get("/api/zones")
def api_get_zones(camera: str = Query(...)):
    s = _store()
    try:
        return s.list_zones(camera)
    finally:
        s.close()


@app.post("/api/zones")
def api_save_zones(payload: ZonePayload):
    s = _store()
    try:
        s.clear_zones(payload.camera)   # editör tam durumu gönderir → değiştir
        for z in payload.zones:
            s.add_zone(payload.camera, z.get("kind", "line"), z.get("name", ""),
                       z.get("points", []), z.get("classes", []),
                       z.get("direction", "AtoB"))
        return {"ok": True, "count": len(payload.zones)}
    finally:
        s.close()


def _saved_intrusions(camera_id: str) -> list[dict]:
    """Kameranın 'intrusion' bölgeleri [{name,points,classes}] — ihlal alanları."""
    out: list[dict] = []
    s = _store()
    try:
        for z in s.list_zones(camera_id):
            if z["kind"] == "intrusion" and len(z["points"] or []) >= 3:
                out.append({"name": z.get("name") or "İhlal alanı",
                            "points": z["points"], "classes": z.get("classes") or []})
    finally:
        s.close()
    return out


def _saved_lines(camera_id: str) -> list[dict]:
    """Kameranın kayıtlı TÜM 'line' bölgelerini [{name,pts,direction}] döndürür."""
    out: list[dict] = []
    s = _store()
    try:
        for z in s.list_zones(camera_id):
            if z["kind"] == "line":
                pts = z["points"] or []
                if len(pts) >= 2:
                    out.append({"name": z.get("name") or "Çizgi", "pts": [pts[0], pts[1]],
                                "direction": z.get("direction") or "AtoB"})
    finally:
        s.close()
    return out


class RunPayload(BaseModel):
    camera: str
    kind: str = "count"   # count | plate | face | analyze
    realtime: bool = True  # dosya kaynağını kamera hızında oynat (bkz. _frame_pusher)


def _webify(video_rel: str) -> None:
    """OpenCV 'mp4v' çıktısını tarayıcı-uyumlu H.264'e çevirir.

    Tarayıcılar mp4v (MPEG-4 Part 2) oynatmaz; H.264 (avc1) gerekir.
    ffmpeg CLI varsa o kullanılır; yoksa PyAV (pip 'av', libx264 bundle'lı) —
    GB10/DGX OS'ta host ffmpeg'i kurulu gelmiyor.
    """
    import shutil
    import subprocess

    name = video_rel.rsplit("/", 1)[-1]
    path = ROOT / cfg.get("paths.output_dir", "output") / name
    if not path.exists():
        return
    tmp = path.with_name(path.stem + "_web.mp4")
    ff = shutil.which("ffmpeg")
    try:
        if ff:
            subprocess.run([ff, "-y", "-i", str(path), "-c:v", "libx264", "-preset", "veryfast",
                            "-pix_fmt", "yuv420p", "-movflags", "+faststart", "-an",
                            "-loglevel", "error", str(tmp)], check=True, timeout=300)
        else:
            _webify_av(path, tmp)
        tmp.replace(path)
    except Exception as e:
        print(f"[server] webify başarısız ({name}): {e}", flush=True)
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError:
                pass


def _webify_av(src, dst) -> None:
    """PyAV ile mp4v → H.264 (faststart, ses yok)."""
    import av

    with av.open(str(src)) as inp, av.open(str(dst), "w", options={"movflags": "+faststart"}) as out:
        ivs = inp.streams.video[0]
        ovs = out.add_stream("libx264", rate=ivs.average_rate or 25)
        ovs.width, ovs.height = ivs.width, ivs.height
        ovs.pix_fmt = "yuv420p"
        ovs.options = {"preset": "veryfast"}
        for frame in inp.decode(ivs):
            out.mux(ovs.encode(frame))
        out.mux(ovs.encode(None))


JOBS: dict[str, dict] = {}
RUN_LOCK = threading.Lock()   # aynı anda tek analiz (SQLite yazma çakışmasını önler)
# Status geçişlerini atomikleştirir: endpoint'in "cancelling" check-and-set'i ile analiz
# thread'inin terminal yazımı (done/error/cancelled) yarışırsa done'un üzerine
# cancelling yazılıp job sonsuza dek dönerdi. Kilit YALNIZ status geçişini sarar.
JOBS_STATE_LOCK = threading.Lock()

# Canlı önizleme: analiz karesi bellekte JPEG olarak tutulur (diske YAZILMAZ — KVKK).
_LIVE_MIN_INTERVAL = 0.05   # sn — UI ~150ms poll'luyor, daha sık encode israf
_LIVE_MAX_W = 960


def _frame_pusher(job: dict, pace: float = 0.0):
    """Analiz modüllerinin on_frame callback'i: annotated kareyi throttle'layıp
    job["frame_jpeg"]'e koyar. UI /api/run/{id}/frame ile çeker.

    pace > 0 ise analiz KAYNAK HIZINA bağlanır (saniye/işlenen kare). GPU dosyayı
    gerçek zamandan ~25x hızlı bitiriyor; önizleme "hızlı ileri sarma" gibi akıp
    sona eriyordu. Test ekranı kamerayı taklit etmeli — canlı kurulumda kareler
    zaten kamera hızında gelir, orada bu bekleme kendiliğinden sıfırdır.
    """
    state = {"t": 0.0, "t0": 0.0, "n": 0}

    def push(frame) -> None:
        if pace > 0:
            if state["t0"] == 0.0:
                state["t0"] = time.monotonic()
            state["n"] += 1
            gecikme = state["t0"] + state["n"] * pace - time.monotonic()
            if gecikme > 0:
                time.sleep(gecikme)
            elif gecikme < -pace:
                # Analiz takvimin GERİSİNE düştü (GPU başka işle meşgul, kare
                # ağır geldi). Mutlak takvime yetişmeye çalışmak, biriken
                # kareleri ardarda basmak demek — ekranda "dondu sonra hızlandı"
                # olarak görünüyordu. Önizleme geri kalabilir, HIZLANAMAZ:
                # başlangıcı şimdiye çekip yeni tempoyu buradan sürdür.
                state["t0"] = time.monotonic() - state["n"] * pace
        # Canlı önizleme hatası analizi ASLA düşürmez (callback modül döngüsünde koşar;
        # istisna fırlarsa writer finalize edilmeden çıkılır)
        try:
            now = time.monotonic()
            if now - state["t"] < _LIVE_MIN_INTERVAL:
                return
            state["t"] = now
            h, w = frame.shape[:2]
            if w > _LIVE_MAX_W:
                frame = cv2.resize(frame, (_LIVE_MAX_W, int(h * _LIVE_MAX_W / w)))
            ok, buf = cv2.imencode(".jpg", frame, [int(cv2.IMWRITE_JPEG_QUALITY), 80])
            if ok:
                job["frame_jpeg"] = buf.tobytes()
                job["frame_seq"] = job.get("frame_seq", 0) + 1
        except Exception:
            pass

    return push


def _pace_seconds(source: str) -> float:
    """İşlenen kare başına düşen kaynak süresi (vid_stride dahil). 0 = tempo yok."""
    if str(source).startswith(("rtsp://", "rtmp://", "http://", "https://")):
        return 0.0   # canlı kaynak zaten kendi hızında akar
    cap = cv2.VideoCapture(source)
    fps = (cap.get(cv2.CAP_PROP_FPS) or 0) if cap.isOpened() else 0
    cap.release()
    if fps <= 0:
        return 0.0
    return max(1, int(cfg.get("detect.vid_stride", 1))) / fps


class _SandboxStore:
    """Test koşusu için olay YUTUCU — veritabanına hiçbir şey yazmaz.

    Test ekranı eskiden gerçek tablolara yazıyor, her koşu öncesi de
    clear_analysis() ile TÜMÜNÜ siliyordu (WHERE'siz DELETE). POC döneminde
    tek yazan test olduğu için sorun değildi; worker sürekli GERÇEK olay
    üretmeye başlayınca felakete dönüştü: operatörün tek test tıklaması tüm
    kameraların olay arşivini (kanıt!) siliyordu — sahada yaşandı, abbey-road
    verisi böyle gitti. Test artık kum havuzunda: sonuçlar job özetinde
    gösterilir, arşive dokunulmaz.
    """

    def add_count_event(self, *a, **k): pass
    def add_plate_event(self, *a, **k): pass
    def add_face_event(self, *a, **k): pass
    def add_alert(self, *a, **k): pass
    def commit(self): pass
    def close(self): pass


def _run_analysis(job_id: str, p: "RunPayload") -> None:
    job = JOBS[job_id]
    cam = _camera(p.camera)
    if not cam:
        with JOBS_STATE_LOCK:
            job.update(status="error", error="Kamera bulunamadı")
        return
    if not RUN_LOCK.acquire(blocking=False):
        job.update(stage="sırada bekliyor (başka analiz çalışıyor)")
        RUN_LOCK.acquire()
    source = cam["source"]; stem = Path(source).stem
    # Canlı HTTP/HLS kaynağı go2rtc RTSP rölesinden alınır (recorder ile aynı
    # gerekçe): ffmpeg ham HLS'i segment segment okur — 6 sn'lik patlama +
    # bekleme. Test önizlemesinde "hızlanıp donuyor" olarak görülen buydu;
    # duvar için -readrate ile çözülmüştü ama test ham adresi açıyordu.
    # Röle ayrıca kaynaktan İKİNCİ bir çekim açılmasını da önler.
    if str(source).startswith(("http://", "https://")):
        go2rtc = (cfg.get("go2rtc.url", "") or "").rstrip("/")
        if go2rtc:
            host = go2rtc.split("//", 1)[-1].split(":")[0] or "localhost"
            source = f"rtsp://{host}:8554/{p.camera}"
    s = None; summary: dict = {}; videos: list[str] = []
    push_frame = _frame_pusher(job, _pace_seconds(source) if p.realtime else 0.0)
    try:
        # Kuyrukta beklerken iptal edildiyse hiç başlama (kilit finally'de bırakılır)
        if job["cancel"].is_set():
            with JOBS_STATE_LOCK:
                job.update(status="cancelled", cancelled=True, stage="iptal edildi", videos=[])
            return
        # store açılışı da try içinde: hata olursa kilit finally'de MUTLAKA bırakılır
        s = _store()          # yalnız OKUMA için (izleme listeleri, kayıtlı çizgiler)
        kum = _SandboxStore() # test olayları arşive DEĞİL buraya
        if p.kind in ("count", "analyze"):
            saved = _saved_lines(p.camera)
            ihlal = _saved_intrusions(p.camera)
            if not saved and not ihlal:
                # Varsayılan orta çizgiyle sessizce saymak yanıltıcı — sayım atlanır,
                # UI kullanıcıyı Bölgeler ekranına yönlendirir
                summary["count"] = {"no_line": True}
            else:
                job.update(stage="Sayım çalışıyor")
                job["count_live"] = {"in": 0, "out": 0, "events": []}
                from .count import run_count
                def _on_count_event(ev):
                    live = job.setdefault("count_live", {"in": 0, "out": 0, "events": []})
                    live["in"] = ev.get("in", live.get("in", 0))
                    live["out"] = ev.get("out", live.get("out", 0))
                    live.setdefault("events", []).append(ev)
                    live["events"] = live["events"][-40:]
                def _on_intrusion(al):
                    live = job.setdefault("count_live", {"in": 0, "out": 0, "events": []})
                    live.setdefault("intrusions", []).append(al)
                    live["intrusions"] = live["intrusions"][-20:]
                s.start_run("count", source)
                res = run_count(source, cfg, save_video=True, store=kum, camera_id=p.camera,
                                lines=saved, on_event=_on_count_event,
                                on_frame=push_frame,
                                should_stop=job["cancel"].is_set,
                                intrusions=ihlal, on_alert=_on_intrusion)
                summary["count"] = {"in": res.in_count, "out": res.out_count,
                                    "frames": res.frames, "lines": res.lines,
                                    "intrusions": res.intrusions}
                videos.append(f"/media/{stem}_count.mp4")
        # İptal edildiyse kalan modüller VE _webify atlanır (yarım videoya dönüşüm israf)
        if job["cancel"].is_set():
            with JOBS_STATE_LOCK:
                job.update(status="cancelled", cancelled=True, stage="iptal edildi", videos=[])
            return
        if p.kind in ("plate", "analyze"):
            job.update(stage="Plaka çalışıyor")
            job["live"] = []   # okundukça canlı eklenir (UI aşağı akıtır)
            from .plate import run_plate
            s.start_run("plate", source)
            res = run_plate(source, cfg, save_video=True, store=kum, camera_id=p.camera,
                            on_read=lambda pl, c, f, t: job["live"].append(
                                {"plate": pl, "conf": round(c, 2) if c else None, "frame": f, "ts": t}),
                            on_frame=push_frame,
                            should_stop=job["cancel"].is_set)
            voted = res.voted or [{"plate": x, "count": 1, "conf": None} for x in res.plates]
            plates = [v["plate"] for v in voted]
            # Uyarı kapısı (ingestor ile aynı kural): tek okumalık eşleşme alarm olmaz
            amr = int(cfg.get("plate.alert_min_reads", 2))
            matches = s.match_plates([v["plate"] for v in voted if v["count"] >= amr])
            # Test alarmı ÜRETMEZ: eşleşme job özetinde gösterilir. Gerçek alarm
            # yalnız worker'dan doğar — demo videodaki plaka, uyarı merkezini
            # sahte kayıtla doldurmasın
            summary["plate"] = {"plates": plates, "total": res.total_reads,
                                "voted": voted, "alerts": matches}
            videos.append(f"/media/{stem}_plate.mp4")
        if job["cancel"].is_set():
            with JOBS_STATE_LOCK:
                job.update(status="cancelled", cancelled=True, stage="iptal edildi", videos=[])
            return
        if p.kind in ("face", "analyze"):
            job.update(stage="Yüz çalışıyor")
            job["face_live"] = {"frames": 0, "raw": 0, "active": 0, "detections": 0, "male": 0, "female": 0, "avg_age": 0.0, "matches": []}
            from .face import run_face
            def _on_face_progress(data):
                job["face_live"] = data
            watch = s.faces_with_embedding()
            s.start_run("face", source)
            res = run_face(source, cfg, save_video=True, store=kum, camera_id=p.camera, watch=watch,
                           on_progress=_on_face_progress, on_frame=push_frame,
                           should_stop=job["cancel"].is_set)
            # yüz eşleşmeleri de özetle sınırlı (üstteki plaka gerekçesi)
            summary["face"] = {"detections": res.detections, "male": res.male, "female": res.female,
                               "avg_age": round(res.avg_age, 1), "alerts": res.matches}
            videos.append(f"/media/{stem}_face.mp4")
        if job["cancel"].is_set():
            with JOBS_STATE_LOCK:
                job.update(status="cancelled", cancelled=True, stage="iptal edildi", videos=[])
            return
        job.update(stage="Video hazırlanıyor")
        for v in videos:
            _webify(v)
        with JOBS_STATE_LOCK:
            job.update(status="done", stage="bitti", summary=summary, videos=videos)
    except Exception as e:  # job hatayı taşır, sunucu çökmez
        with JOBS_STATE_LOCK:
            job.update(status="error", error=str(e))
    finally:
        # KVKK veri minimizasyonu: analiz bitince son annotated kare bellekte kalmaz
        job.pop("frame_jpeg", None)
        if s is not None:
            s.close()
        RUN_LOCK.release()


@app.post("/api/run")
def api_run(p: RunPayload):
    # Yeni koşu eski koşuyu bekletmez: çalışan tüm job'lar iptale çekilir
    # (nihai "cancelled" durumunu analiz thread'i yazar). Check-and-set kilit
    # altında: thread'in az önce yazdığı terminal durum ezilmez.
    with JOBS_STATE_LOCK:
        for j in list(JOBS.values()):
            if j.get("status") == "running":
                ev = j.get("cancel")
                if ev is not None:
                    ev.set()
                j["status"] = "cancelling"
    job_id = uuid.uuid4().hex[:12]
    while len(JOBS) > 50:   # eski job kayıtları birikmesin
        JOBS.pop(next(iter(JOBS)))
    JOBS[job_id] = {"status": "running", "stage": "başlıyor", "summary": {}, "videos": [],
                    "cancel": threading.Event()}
    threading.Thread(target=_run_analysis, args=(job_id, p), daemon=True).start()
    return {"job_id": job_id}


@app.get("/api/run/{job_id}")
def api_run_status(job_id: str):
    j = JOBS.get(job_id)
    if not j:
        raise HTTPException(404, "job bulunamadı")
    # dict(j): analiz thread'i eşzamanlı anahtar ekler — kopya GIL altında atomik,
    # kilitisiz iterasyon RuntimeError'ı önlenir. frame_jpeg ham bytes, cancel bir
    # threading.Event — ikisi de JSON'a girmez.
    j = dict(j)
    j.pop("frame_jpeg", None)
    j.pop("cancel", None)
    return {"job_id": job_id, **j}


@app.delete("/api/run/{job_id}")
def api_run_cancel(job_id: str):
    """Koşan analizi iptal eder; nihai 'cancelled' durumunu analiz thread'i yazar."""
    j = JOBS.get(job_id)
    if not j:
        raise HTTPException(404, "job bulunamadı")
    ev = j.get("cancel")
    if ev is not None:
        ev.set()
    # Check-and-set kilit altında: thread'in terminal yazımının üzerine yazılmaz
    with JOBS_STATE_LOCK:
        if j.get("status") == "running":
            j["status"] = "cancelling"
    return {"ok": True}


@app.get("/api/run/{job_id}/frame")
def api_run_frame(job_id: str):
    """Analiz sürerken son annotated kare (JPEG, yalnız bellek — diske yazılmaz)."""
    j = JOBS.get(job_id)
    if not j:
        raise HTTPException(404, "job bulunamadı")
    data = j.get("frame_jpeg")
    if not data:
        return Response(status_code=204)
    return Response(content=data, media_type="image/jpeg",
                    headers={"Cache-Control": "no-store"})


@app.get("/api/events")
def api_events(limit: int = Query(50, ge=1, le=500), tur: str = "", kamera: str = ""):
    # sınırsız int SQLite'ı taşırıp 500 döndürüyordu (Schemathesis bulgusu) — 422'ye bağlanır
    if tur and tur not in ("count", "plate", "face"):
        raise HTTPException(422, "tur: count | plate | face")
    s = _store()
    try:
        olaylar = s.recent_events(limit, tur, kamera)
    finally:
        s.close()
    # Yabancı plaka etiketi: TR formatı yapısal doğrulamadan geçer, yabancı
    # plaka yalnız güven eşiğinden. Operatör aradaki farkı GÖRMELİ — ikisini
    # aynı göstermek, olmayan bir güvence vermek olur.
    from .plate import plaka_turu
    for o in olaylar:
        if o.get("type") == "plate" and o.get("detail"):
            o["plaka_tur"] = plaka_turu(str(o["detail"]).split()[0])
    return olaylar


@app.get("/api/events/summary")
def api_events_summary(hours: int = Query(24, ge=1, le=720)):
    """Kamera başına olay dökümü — "hangi kamerada ne oluyor" tek bakışta.

    Olay akışını kamera kamera taramak yerine operatör önce özete bakar;
    dikkat isteyen kamerayı oradan seçer.
    """
    from datetime import timedelta

    sinir = datetime.now(timezone.utc) - timedelta(hours=hours)
    s = _store()
    try:
        olaylar = s.recent_events(20000)
        bekleyen = s.recent_alerts(500, pending_only=True)
    finally:
        s.close()
    ozet: dict[str, dict] = {}
    for e in olaylar:
        try:
            t = datetime.fromisoformat(str(e["time"]).replace(" ", "T"))
            if t.tzinfo is None:
                t = t.replace(tzinfo=timezone.utc)
            if t < sinir:
                continue
        except ValueError:
            pass
        k = ozet.setdefault(e["camera_id"], {"camera_id": e["camera_id"], "count": 0,
                                             "plate": 0, "face": 0, "alerts": 0,
                                             "last": None})
        k["count"] += 1
        if e["type"] in k:
            k[e["type"]] += 1
        if k["last"] is None:
            k["last"] = str(e["time"])   # liste zaten zaman DESC
    for a in bekleyen:
        k = ozet.setdefault(a["camera_id"] or "?", {"camera_id": a["camera_id"] or "?",
                                                    "count": 0, "plate": 0, "face": 0,
                                                    "alerts": 0, "last": None})
        k["alerts"] += 1
    return {"hours": hours,
            "cameras": sorted(ozet.values(), key=lambda x: (-x["alerts"], -x["count"]))}


@app.get("/api/alerts")
def api_alerts(limit: int = Query(20, ge=1, le=500), pending: bool = False):
    s = _store()
    try:
        return s.recent_alerts(limit, pending_only=pending)
    finally:
        s.close()


@app.post("/api/alerts/{alert_id}/ack")
def api_ack_alert(alert_id: int, request: Request):
    """Uyarıyı kabul eder — operatör gördü, işlem yapıldı.

    Kabul edilen uyarı bekleyen listesinden ve menü rozetinden düşer; kayıt silinmez
    (kim ne zaman kabul etti bilgisi denetim izi olarak kalır).
    """
    kim = (_istek_kullanicisi(request) or {}).get("ad") or "operatör"
    s = _store()
    try:
        if not s.ack_alert(alert_id, by=kim):
            raise HTTPException(404, "Uyarı bulunamadı veya zaten kabul edilmiş")
        return {"ok": True}
    finally:
        s.close()


@app.get("/api/counts")
def api_counts():
    # active: şu an koşan analiz var mı — UI sayaç kutularını yalnız analiz
    # sürerken gösterir (ekran ilk açılışta eski toplamlarla dolmasın).
    # "cancelling" de aktif sayılır: iptal isteği işlenene dek analiz hâlâ koşuyor.
    active = any(j.get("status") in ("running", "cancelling") for j in list(JOBS.values()))
    s = _store()
    try:
        return {"active": active, "rows": s.count_totals()}
    finally:
        s.close()


class WatchPayload(BaseModel):
    kind: str           # plate | face
    value: str          # plaka metni veya kişi adı
    label: str = ""
    list_type: str = "blacklist"
    camera: str = ""    # face: bu kameranın karesinden embedding çıkar (enroll)


@app.get("/api/lists")
def api_lists(kind: str = Query("plate")):
    s = _store()
    try:
        return s.list_watch(kind)
    finally:
        s.close()


@app.post("/api/lists")
def api_add_watch(p: WatchPayload):
    s = _store()
    try:
        if p.kind == "plate":
            s.add_watch_plate(p.value, p.label, p.list_type)
            return {"ok": True}
        # yüz: kamera verildiyse o kareden embedding çıkar (enroll)
        emb = None
        if p.camera:
            cam = _camera(p.camera)
            if cam:
                from .face import embed_largest_face
                emb = embed_largest_face(cam["source"], cfg)
        s.add_watch_face(p.value, p.label, p.list_type, embedding=emb)
        return {"ok": True, "enrolled": emb is not None}
    finally:
        s.close()


@app.delete("/api/lists/{kind}/{row_id}")
def api_del_watch(kind: str, row_id: int):
    s = _store()
    try:
        s.delete_watch(kind, row_id)
        return {"ok": True}
    finally:
        s.close()


@app.get("/m")
@app.get("/m/")
def mobil():
    """Mobil arayüz (PWA) — masaüstü paneliyle AYNI API'leri kullanır.

    Ayrı uygulama/mağaza süreci yok: telefondan /m açılır, "Ana ekrana ekle"
    ile kurulur. Erişim anahtarı ?token= ile bir kez verilir.
    """
    return FileResponse(WEB_DIR / "mobil.html", headers={"Cache-Control": "no-cache"})


def _yerel_ip() -> str:
    """Telefonun ulaşabileceği yerel IP — 127.0.0.1 telefonda İŞE YARAMAZ.

    Dışa bağlantı açmadan (UDP soketi paket göndermez) işletim sisteminin
    varsayılan arayüzünü sorar.
    """
    import socket

    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(("8.8.8.8", 80))
        return s.getsockname()[0]
    except Exception:
        return "127.0.0.1"
    finally:
        s.close()


@app.get("/api/mobil-qr")
def api_mobil_qr(host: str = ""):
    """Mobil arayüzü telefonda açan QR kod (SVG) + adres.

    Operatör IP ve uzun erişim anahtarını elle yazmaz: QR'ı okutur, mobil
    doğrudan yetkili şekilde açılır. Adres olarak MAKİNENİN YEREL IP'si
    kullanılır — panel 127.0.0.1'de açık olsa bile telefon oraya ulaşamaz.
    """
    import segno

    ip = (host or _yerel_ip()).strip()
    port = int(cfg.get("server.port", 8000))
    adres = f"http://{ip}:{port}/m"
    if API_TOKEN:
        adres += f"?token={API_TOKEN}"
    qr = segno.make(adres, error="m")
    svg = qr.svg_inline(scale=6, dark="#0e1521", light="#ffffff", border=2)
    return {"adres": adres, "ip": ip, "port": port, "svg": svg,
            "uyari": ("Bu adres yalnız aynı yerel ağdan çalışır. Dışarıdan erişim için "
                      "VPN gerekir; erişim anahtarı şifresiz HTTP üzerinden gider.")}


@app.get("/m/manifest.json")
def mobil_manifest():
    """PWA tanımı — telefona kurulabilmesi ve tam ekran açılması için."""
    return {
        "name": "AurasVision", "short_name": "AurasVision",
        "description": "Güvenlik kamerası izleme ve alarm merkezi",
        "start_url": "/m", "scope": "/m", "display": "standalone",
        "orientation": "portrait", "background_color": "#070b12",
        "theme_color": "#070b12", "lang": "tr",
        # Simge gömülü SVG: dış dosya bağımlılığı yok (çevrimdışı kurulum)
        "icons": [{
            "src": ("data:image/svg+xml;utf8,<svg xmlns='http://www.w3.org/2000/svg' "
                    "viewBox='0 0 64 64'><rect width='64' height='64' rx='14' fill='%23070b12'/>"
                    "<g fill='none' stroke='%2338bdf8' stroke-width='3' stroke-linecap='round'>"
                    "<circle cx='32' cy='32' r='18'/><path d='M32 8v10M32 46v10M8 32h10M46 32h10'/></g>"
                    "<circle cx='32' cy='32' r='6' fill='%2338bdf8'/></svg>"),
            "sizes": "any", "type": "image/svg+xml", "purpose": "any"}],
    }


@app.get("/")
def index():
    # no-cache: sürüm güncellendiğinde operatör eski arayüzle kalmasın
    # (tarayıcı yine 304 ile doğrular — bant genişliği israfı değil, bayat UI yok)
    return FileResponse(WEB_DIR / "index.html",
                        headers={"Cache-Control": "no-cache"})


if WEB_DIR.exists():
    app.mount("/static", StaticFiles(directory=WEB_DIR), name="static")

_OUT = ROOT / (cfg.get("paths.output_dir", "output"))
_OUT.mkdir(parents=True, exist_ok=True)
app.mount("/media", StaticFiles(directory=_OUT), name="media")


class _QuietPolls(logging.Filter):
    """UI'ın saniyede birkaç kez tekrarlayan poll'ları access log'u boğmasın."""

    def filter(self, record: logging.LogRecord) -> bool:
        msg = record.getMessage()
        return not ("GET /api/counts" in msg or "GET /api/run/" in msg)


def main() -> None:
    import uvicorn
    from .gunluk import kur as gunluk_kur
    gunluk_kur("sunucu", cfg)

    logging.getLogger("uvicorn.access").addFilter(_QuietPolls())
    _sync_go2rtc()
    host = cfg.get("server.host", "127.0.0.1")
    port = int(cfg.get("server.port", 8000))
    uvicorn.run(app, host=host, port=port)


if __name__ == "__main__":
    main()
